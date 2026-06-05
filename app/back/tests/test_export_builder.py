from datetime import datetime, timezone

import openpyxl
import pytest

from src.engine.export_builder import (
    build_fleet_detail_workbook,
    build_fleets_list_workbook,
    build_transactions_list_workbook,
)
from src.models.transaction import Transaction
from src.models.user import User, UserRole


def _sample_transaction(transaction_id: int = 1) -> Transaction:
    return Transaction(
        id=transaction_id,
        user_id=10,
        vehicle_id=20,
        organization_id=1,
        plate="ABC1D23",
        context="pedagio",
        uf="SP",
        elapsed_time_sec=30,
        is_digital=True,
        co2_avoided_kg=1.5,
        fuel_saved_liters=0.2,
        time_saved_sec=120,
        financial_savings_brl=5.5,
        water_saved_liters=0.1,
        parameters_snapshot={},
        created_at=datetime(2025, 6, 1, 12, 0, tzinfo=timezone.utc),
    )


def test_build_fleets_list_workbook_has_expected_sheet():
    buffer = build_fleets_list_workbook(
        [
            {
                "id": 1,
                "name": "Frota A",
                "organization_id": 2,
                "vehicle_count": 3,
                "created_at": datetime(2025, 1, 1, tzinfo=timezone.utc),
            }
        ]
    )
    wb = openpyxl.load_workbook(buffer)
    assert wb.sheetnames == ["Frotas"]
    assert wb["Frotas"]["A4"].value == "ID"
    assert wb["Frotas"]["B5"].value == "Frota A"


def test_build_fleet_detail_workbook_has_expected_tabs_and_banner():
    fleet = {
        "id": 1,
        "name": "Frota A",
        "organization_id": 2,
        "created_at": datetime(2025, 1, 1, tzinfo=timezone.utc),
    }
    summary = {
        "vehicle_count": 1,
        "driver_count": 1,
        "transaction_count": 1,
        "co2_total_kg": 1.0,
        "fuel_total_liters": 0.5,
        "total_savings_brl": 10.0,
        "paper_saved_meters": 2.0,
    }
    vehicle = type(
        "VehicleStub",
        (),
        {
            "id": 1,
            "id_tag": "TAG001",
            "license_plate": "ABC1D23",
            "model": "Sedan",
            "fuel_type": "gasolina_c",
            "category": "leve",
            "organization_id": 2,
        },
    )()
    driver = User(
        id=5,
        name="João",
        email="joao@example.com",
        password_hash="hash",
        role=UserRole.motorista,
        organization_id=2,
    )
    buffer = build_fleet_detail_workbook(
        fleet,
        summary,
        [vehicle],
        [driver],
        [_sample_transaction()],
    )
    wb = openpyxl.load_workbook(buffer)
    assert wb.sheetnames == ["Frota", "Veículos", "Motoristas", "Passagens"]
    assert "Template: taggy_ecoscore_frota" in str(wb["Frota"]["A3"].value)


def test_build_transactions_list_workbook_columns():
    buffer = build_transactions_list_workbook([_sample_transaction()])
    wb = openpyxl.load_workbook(buffer)
    ws = wb["Passagens"]
    assert ws["A4"].value == "ID"
    assert ws["B4"].value == "Placa"
    assert ws["B5"].value == "ABC1D23"


def test_build_fleets_list_workbook_requires_openpyxl(monkeypatch):
    monkeypatch.setattr("src.engine.export_builder._HAS_OPENPYXL", False)
    with pytest.raises(RuntimeError, match="openpyxl"):
        build_fleets_list_workbook([])
