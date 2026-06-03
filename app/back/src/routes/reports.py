import io
from datetime import date
from typing import Literal

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession

from src.database.connection import get_db
from src.engine import CalcEngine, CalcEngineError, TransactionOrchestrator
from src.engine.report_builder import build_audit_workbook
from src.repositories.transaction_repository import TransactionRepository
from src.services.technical_specs import get_all_specs
from src.services.vehicle_lookup_service import resolve_vehicle_from_plate

router = APIRouter(prefix="/reports", tags=["Reports"])


@router.get("/export")
async def export_transactions(
    user_id: int = Query(...),
    from_date: date | None = Query(default=None),
    to_date: date | None = Query(default=None),
    db: AsyncSession = Depends(get_db),
) -> StreamingResponse:
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError as exc:
        raise RuntimeError("openpyxl não está instalado.") from exc

    repo = TransactionRepository(db)
    transactions = await repo.get_by_user_in_range(user_id, from_date, to_date)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Relatório de Passagens"

    headers = [
        "ID",
        "Placa",
        "Contexto",
        "UF",
        "Data/Hora",
        "CO₂ Evitado (kg)",
        "Combustível Economizado (L)",
        "Tempo Economizado (min)",
        "Economia Financeira (R$)",
        "Água Economizada (L)",
    ]

    header_fill = PatternFill(start_color="1A1A2E", end_color="1A1A2E", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    context_labels = {"pedagio": "Pedágio", "estacionamento": "Estacionamento"}

    for row_idx, t in enumerate(transactions, start=2):
        ws.cell(row=row_idx, column=1, value=t.id)
        ws.cell(row=row_idx, column=2, value=t.plate or "")
        ws.cell(row=row_idx, column=3, value=context_labels.get(t.context, t.context))
        ws.cell(row=row_idx, column=4, value=t.uf or "")
        ws.cell(row=row_idx, column=5, value=t.created_at.strftime("%d/%m/%Y %H:%M:%S") if t.created_at else "")
        ws.cell(row=row_idx, column=6, value=round(t.co2_avoided_kg or 0, 4))
        ws.cell(row=row_idx, column=7, value=round(t.fuel_saved_liters or 0, 4))
        ws.cell(row=row_idx, column=8, value=round((t.time_saved_sec or 0) / 60, 2))
        ws.cell(row=row_idx, column=9, value=round(t.financial_savings_brl or 0, 2))
        ws.cell(row=row_idx, column=10, value=round(t.water_saved_liters or 0, 4))

    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max(max_len + 4, 14)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"relatorio_passagens_usuario_{user_id}.xlsx"

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/calculadora.xlsx")
async def export_calculadora(
    plate: str = Query(default="DEMO0001", max_length=10),
    elapsed_time: int = Query(default=30, ge=0),
    context: Literal["pedagio", "estacionamento"] = Query(default="pedagio"),
    uf: str = Query(default="SP", min_length=2, max_length=2),
    is_digital: bool = Query(default=True),
    fleet_size: int = Query(default=1, ge=1, le=100000),
    fuel_type: str | None = Query(default=None),
    category: str | None = Query(default=None),
    db: AsyncSession = Depends(get_db),
) -> StreamingResponse:
    """
    Gera planilha auditável com 5 sheets:
    Premissas | Passo a Passo | Comparação | Sensibilidade | Escala
    """
    try:
        specs = await get_all_specs(db)
    except CalcEngineError as e:
        raise HTTPException(status_code=422, detail=str(e)) from e

    # Resolve vehicle data
    if fuel_type and category:
        vehicle = {"category": category, "fuel_type": fuel_type, "model": ""}
    else:
        lookup = await resolve_vehicle_from_plate(plate)
        if lookup["error"] or lookup["vehicle"] is None:
            # Fallback to sensible demo values
            vehicle = {"category": "leve", "fuel_type": "gasolina_c", "model": "Veículo demo"}
        else:
            vehicle = lookup["vehicle"]

    payload = {
        "plate": plate.upper(),
        "elapsed_time": elapsed_time,
        "context": context,
        "uf_passagem": uf.upper(),
        "is_digital": is_digital,
        "vehicle": vehicle,
    }

    try:
        engine = CalcEngine(specs)
        orchestrator = TransactionOrchestrator(engine)
        result = orchestrator.handle_tag_event(payload)
    except CalcEngineError as e:
        raise HTTPException(status_code=422, detail=str(e)) from e

    params = {
        "plate": plate.upper(),
        "elapsed_time": elapsed_time,
        "context": context,
        "uf": uf.upper(),
        "is_digital": is_digital,
    }

    try:
        buffer = build_audit_workbook(result, specs, vehicle, params, fleet_size=fleet_size)
    except RuntimeError as e:
        raise HTTPException(status_code=500, detail=str(e)) from e

    filename = f"ecoscore_calculo_{plate.upper()}_{context}_{uf.upper()}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
