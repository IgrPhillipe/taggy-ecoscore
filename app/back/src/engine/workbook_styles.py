"""Estilos compartilhados das planilhas Taggy EcoScore (exports e templates)."""

from __future__ import annotations

from datetime import datetime

from src.constants.workbook_exports import workbook_export_version_label, workbook_template_label

try:
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.worksheet import Worksheet
except ImportError:
    Worksheet = object  # type: ignore[misc, assignment]

_C_HEADER_BG = "1A3A4A"
_C_HEADER_FG = "FFFFFF"
_C_ALT_BG = "F5F8FA"
_C_BORDER = "CCCCCC"


def font_title() -> Font:
    return Font(bold=True, name="Calibri", size=14)


def font_small() -> Font:
    return Font(name="Calibri", size=9, color="555555")


def font_body(*, bold: bool = False, size: int = 10) -> Font:
    return Font(bold=bold, name="Calibri", size=size, color="000000")


def fill_header() -> PatternFill:
    return PatternFill("solid", fgColor=_C_HEADER_BG)


def fill_alt() -> PatternFill:
    return PatternFill("solid", fgColor=_C_ALT_BG)


def thin_border() -> Border:
    side = Side(style="thin", color=_C_BORDER)
    return Border(left=side, right=side, top=side, bottom=side)


def write_sheet_banner(
    ws: Worksheet,
    *,
    title: str,
    subtitle: str,
    template_slug: str,
    merge_cols: int,
) -> int:
    """Escreve cabeçalho visual (3 linhas). Retorna a linha do header da tabela."""
    end_col = get_column_letter(merge_cols)
    ws.merge_cells(f"A1:{end_col}1")
    ws["A1"] = title
    ws["A1"].font = font_title()
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 32

    ws.merge_cells(f"A2:{end_col}2")
    ws["A2"] = subtitle
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="555555")
    ws["A2"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[2].height = 28

    exported_at = datetime.now().strftime("%d/%m/%Y %H:%M")
    ws.merge_cells(f"A3:{end_col}3")
    ws["A3"] = (
        f"Template: {workbook_template_label(template_slug)} · "
        f"Versão: {workbook_export_version_label()} · Exportado em: {exported_at}"
    )
    ws["A3"].font = font_small()
    ws["A3"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[3].height = 18
    return 4


def write_header_row(ws: Worksheet, row: int, headers: list[str]) -> None:
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = Font(bold=True, color=_C_HEADER_FG, name="Calibri", size=11)
        cell.fill = fill_header()
        cell.border = thin_border()
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def write_data_cell(
    ws: Worksheet,
    row: int,
    col: int,
    value,
    *,
    alt: bool = False,
    bold: bool = False,
) -> None:
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = font_body(bold=bold)
    cell.border = thin_border()
    cell.alignment = Alignment(wrap_text=True, vertical="top")
    if alt:
        cell.fill = fill_alt()


def set_col_widths(ws: Worksheet, widths: list[int]) -> None:
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width


def autosize_columns(ws: Worksheet, min_width: int = 14) -> None:
    for col_idx, col in enumerate(ws.columns, start=1):
        values = [cell.value for cell in col if getattr(cell, "value", None) is not None]
        if not values:
            continue
        max_len = max(len(str(value)) for value in values)
        ws.column_dimensions[get_column_letter(col_idx)].width = max(max_len + 4, min_width)
