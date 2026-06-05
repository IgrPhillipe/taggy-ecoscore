"""Builders de planilhas XLSX para export de listagens e detalhes."""

from __future__ import annotations

import io
from collections.abc import Sequence
from datetime import datetime
from typing import Any

from src.constants.workbook_exports import (
    FROTA_SLUG,
    FROTAS_SLUG,
    MOTORISTA_SLUG,
    MOTORISTAS_SLUG,
    PASSAGEM_SLUG,
    PASSAGENS_SLUG,
    VEICULO_SLUG,
    VEICULOS_SLUG,
)
from src.engine.report_builder import build_audit_workbook
from src.engine.workbook_styles import (
    autosize_columns,
    set_col_widths,
    write_data_cell,
    write_header_row,
    write_sheet_banner,
)
from src.models.transaction import Transaction
from src.models.user import User
from src.models.vehicle import Vehicle

try:
    import openpyxl
    from openpyxl.worksheet.worksheet import Worksheet

    _HAS_OPENPYXL = True
except ImportError:
    _HAS_OPENPYXL = False
    Worksheet = object  # type: ignore[misc, assignment]

_CONTEXT_LABELS = {"pedagio": "Pedágio", "estacionamento": "Estacionamento"}

_TRANSACTION_HEADERS = [
    "ID",
    "Placa",
    "Contexto",
    "UF",
    "Digital",
    "Tempo (s)",
    "CO₂ Evitado (kg)",
    "Combustível (L)",
    "Tempo Econ. (min)",
    "Economia (R$)",
    "Água (L)",
    "User ID",
    "Veículo ID",
    "Org ID",
    "Data/Hora",
]


def _require_openpyxl() -> None:
    if not _HAS_OPENPYXL:
        raise RuntimeError("openpyxl não está instalado. Execute: uv add openpyxl")


def _format_dt(value: datetime | None) -> str:
    if value is None:
        return ""
    return value.strftime("%d/%m/%Y %H:%M:%S")


def _transaction_row(t: Transaction) -> list[Any]:
    return [
        t.id,
        t.plate or "",
        _CONTEXT_LABELS.get(t.context, t.context),
        t.uf or "",
        "Sim" if t.is_digital else "Não",
        round(t.elapsed_time_sec or 0, 0),
        round(t.co2_avoided_kg or 0, 4),
        round(t.fuel_saved_liters or 0, 4),
        round((t.time_saved_sec or 0) / 60, 2),
        round(t.financial_savings_brl or 0, 2),
        round(t.water_saved_liters or 0, 4),
        t.user_id or "",
        t.vehicle_id or "",
        t.organization_id or "",
        _format_dt(t.created_at),
    ]


def _write_table_sheet(
    ws: Worksheet,
    *,
    title: str,
    subtitle: str,
    template_slug: str,
    headers: list[str],
    rows: Sequence[Sequence[Any]],
) -> None:
    header_row = write_sheet_banner(
        ws,
        title=title,
        subtitle=subtitle,
        template_slug=template_slug,
        merge_cols=len(headers),
    )
    write_header_row(ws, header_row, headers)
    for row_idx, row in enumerate(rows, start=header_row + 1):
        alt = row_idx % 2 == 0
        for col_idx, value in enumerate(row, start=1):
            write_data_cell(ws, row_idx, col_idx, value, alt=alt)
    autosize_columns(ws)
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1).coordinate


def _write_key_value_sheet(
    ws: Worksheet,
    *,
    title: str,
    subtitle: str,
    template_slug: str,
    pairs: list[tuple[str, Any]],
) -> None:
    header_row = write_sheet_banner(
        ws,
        title=title,
        subtitle=subtitle,
        template_slug=template_slug,
        merge_cols=2,
    )
    write_header_row(ws, header_row, ["Campo", "Valor"])
    for row_idx, (label, value) in enumerate(pairs, start=header_row + 1):
        alt = row_idx % 2 == 0
        write_data_cell(ws, row_idx, 1, label, alt=alt, bold=True)
        write_data_cell(ws, row_idx, 2, value if value is not None else "", alt=alt)
    set_col_widths(ws, [28, 40])
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1).coordinate


def _save_workbook(wb: openpyxl.Workbook) -> io.BytesIO:
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def build_fleets_list_workbook(fleets: list[dict[str, Any]]) -> io.BytesIO:
    _require_openpyxl()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Frotas"
    _write_table_sheet(
        ws,
        title="Relatório de Frotas — Taggy EcoScore",
        subtitle="Listagem exportada com os filtros aplicados na plataforma.",
        template_slug=FROTAS_SLUG,
        headers=["ID", "Nome", "Organização ID", "Qtd. Veículos", "Criado em"],
        rows=[
            [
                f["id"],
                f["name"],
                f["organization_id"],
                f.get("vehicle_count", 0),
                _format_dt(f.get("created_at")),
            ]
            for f in fleets
        ],
    )
    return _save_workbook(wb)


def build_fleet_detail_workbook(
    fleet: dict[str, Any],
    summary: dict[str, Any],
    vehicles: list[Vehicle],
    drivers: list[User],
    transactions: list[Transaction],
) -> io.BytesIO:
    _require_openpyxl()
    wb = openpyxl.Workbook()

    ws_fleet = wb.active
    ws_fleet.title = "Frota"
    _write_key_value_sheet(
        ws_fleet,
        title=f"Frota #{fleet['id']} — {fleet['name']}",
        subtitle="Resumo consolidado da frota, incluindo KPIs de impacto.",
        template_slug=FROTA_SLUG,
        pairs=[
            ("ID", fleet["id"]),
            ("Nome", fleet["name"]),
            ("Organização ID", fleet["organization_id"]),
            ("Criado em", _format_dt(fleet.get("created_at"))),
            ("Veículos", summary.get("vehicle_count")),
            ("Motoristas", summary.get("driver_count")),
            ("Passagens", summary.get("transaction_count")),
            ("CO₂ total (kg)", round(summary.get("co2_total_kg") or 0, 3)),
            ("Combustível total (L)", round(summary.get("fuel_total_liters") or 0, 3)),
            ("Economia total (R$)", round(summary.get("total_savings_brl") or 0, 2)),
            ("Papel economizado (m)", round(summary.get("paper_saved_meters") or 0, 2)),
        ],
    )

    ws_vehicles = wb.create_sheet("Veículos")
    _write_table_sheet(
        ws_vehicles,
        title="Veículos da Frota",
        subtitle="Veículos vinculados à frota no momento da exportação.",
        template_slug=FROTA_SLUG,
        headers=["ID", "TAG ID", "Placa", "Modelo", "Combustível", "Categoria", "Org ID"],
        rows=[
            [
                v.id,
                v.id_tag,
                v.license_plate,
                v.model,
                v.fuel_type,
                v.category,
                v.organization_id or "",
            ]
            for v in vehicles
        ],
    )

    ws_drivers = wb.create_sheet("Motoristas")
    _write_table_sheet(
        ws_drivers,
        title="Motoristas da Frota",
        subtitle="Motoristas vinculados à frota no momento da exportação.",
        template_slug=FROTA_SLUG,
        headers=["ID", "Nome", "E-mail", "Função", "Org ID"],
        rows=[
            [
                u.id,
                u.name,
                u.email,
                u.role.value if hasattr(u.role, "value") else u.role,
                u.organization_id or "",
            ]
            for u in drivers
        ],
    )

    ws_tx = wb.create_sheet("Passagens")
    _write_table_sheet(
        ws_tx,
        title="Passagens da Frota",
        subtitle="Histórico de passagens associadas aos veículos desta frota.",
        template_slug=FROTA_SLUG,
        headers=_TRANSACTION_HEADERS,
        rows=[_transaction_row(t) for t in transactions],
    )

    return _save_workbook(wb)


def build_vehicles_list_workbook(vehicles: list[Vehicle]) -> io.BytesIO:
    _require_openpyxl()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Veículos"
    _write_table_sheet(
        ws,
        title="Relatório de Veículos — Taggy EcoScore",
        subtitle="Listagem exportada com os filtros aplicados na plataforma.",
        template_slug=VEICULOS_SLUG,
        headers=[
            "ID",
            "TAG ID",
            "Placa",
            "Modelo",
            "Combustível",
            "Categoria",
            "Org ID",
            "Frota ID",
        ],
        rows=[
            [
                v.id,
                v.id_tag,
                v.license_plate,
                v.model,
                v.fuel_type,
                v.category,
                v.organization_id or "",
                v.fleet_id or "",
            ]
            for v in vehicles
        ],
    )
    return _save_workbook(wb)


def build_vehicle_detail_workbook(
    vehicle: Vehicle,
    summary: dict[str, Any],
    transactions: list[Transaction],
) -> io.BytesIO:
    _require_openpyxl()
    wb = openpyxl.Workbook()

    ws_info = wb.active
    ws_info.title = "Veículo"
    _write_key_value_sheet(
        ws_info,
        title=f"Veículo #{vehicle.id} — {vehicle.license_plate}",
        subtitle="Dados cadastrais e KPIs consolidados do veículo.",
        template_slug=VEICULO_SLUG,
        pairs=[
            ("ID", vehicle.id),
            ("TAG ID", vehicle.id_tag),
            ("Placa", vehicle.license_plate),
            ("Modelo", vehicle.model),
            ("Combustível", vehicle.fuel_type),
            ("Categoria", vehicle.category),
            ("Org ID", vehicle.organization_id),
            ("Frota ID", vehicle.fleet_id),
            ("Passagens", summary.get("transaction_count")),
            ("CO₂ total (kg)", round(summary.get("co2_total_kg") or 0, 3)),
            ("Combustível total (L)", round(summary.get("fuel_total_liters") or 0, 3)),
            ("Economia total (R$)", round(summary.get("financial_total_brl") or 0, 2)),
            ("Tempo total (h)", round((summary.get("time_total_sec") or 0) / 3600, 2)),
            ("Papel economizado (m)", round(summary.get("paper_saved_meters") or 0, 2)),
        ],
    )

    ws_tx = wb.create_sheet("Passagens")
    _write_table_sheet(
        ws_tx,
        title=f"Passagens — {vehicle.license_plate}",
        subtitle="Histórico de passagens do veículo conforme filtros aplicados.",
        template_slug=VEICULO_SLUG,
        headers=_TRANSACTION_HEADERS,
        rows=[_transaction_row(t) for t in transactions],
    )

    return _save_workbook(wb)


def build_drivers_list_workbook(
    drivers: list[User],
    plates_by_user_id: dict[int, str | None] | None = None,
) -> io.BytesIO:
    _require_openpyxl()
    plates = plates_by_user_id or {}
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Motoristas"
    _write_table_sheet(
        ws,
        title="Relatório de Motoristas — Taggy EcoScore",
        subtitle="Listagem exportada com os filtros aplicados na plataforma.",
        template_slug=MOTORISTAS_SLUG,
        headers=["ID", "Nome", "E-mail", "Placa", "Org ID"],
        rows=[
            [
                u.id,
                u.name,
                u.email,
                plates.get(u.id or 0) or "",
                u.organization_id or "",
            ]
            for u in drivers
        ],
    )
    return _save_workbook(wb)


def build_driver_detail_workbook(
    driver: User,
    stats: dict[str, Any],
    transactions: list[Transaction],
) -> io.BytesIO:
    _require_openpyxl()
    wb = openpyxl.Workbook()

    ws_info = wb.active
    ws_info.title = "Motorista"
    _write_key_value_sheet(
        ws_info,
        title=f"Motorista #{driver.id} — {driver.name}",
        subtitle="Dados cadastrais e KPIs consolidados do motorista.",
        template_slug=MOTORISTA_SLUG,
        pairs=[
            ("ID", driver.id),
            ("Nome", driver.name),
            ("E-mail", driver.email),
            ("Função", driver.role.value if hasattr(driver.role, "value") else driver.role),
            ("Org ID", driver.organization_id),
            ("Passagens", stats.get("transactions_count")),
            ("CO₂ total (kg)", round(stats.get("co2_total_kg") or 0, 3)),
            ("Combustível total (L)", round(stats.get("fuel_total_liters") or 0, 3)),
            ("Economia total (R$)", round(stats.get("financial_total_brl") or 0, 2)),
            ("Tempo economizado (h)", round((stats.get("total_time_saved_sec") or 0) / 3600, 2)),
            ("Papel economizado (m)", round(stats.get("paper_saved_meters") or 0, 2)),
        ],
    )

    ws_tx = wb.create_sheet("Passagens")
    _write_table_sheet(
        ws_tx,
        title=f"Passagens — {driver.name}",
        subtitle="Histórico de passagens do motorista conforme filtros aplicados.",
        template_slug=MOTORISTA_SLUG,
        headers=_TRANSACTION_HEADERS,
        rows=[_transaction_row(t) for t in transactions],
    )

    return _save_workbook(wb)


def build_transactions_list_workbook(transactions: list[Transaction]) -> io.BytesIO:
    _require_openpyxl()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Passagens"
    _write_table_sheet(
        ws,
        title="Relatório de Passagens — Taggy EcoScore",
        subtitle="Listagem exportada com os filtros aplicados na plataforma.",
        template_slug=PASSAGENS_SLUG,
        headers=_TRANSACTION_HEADERS,
        rows=[_transaction_row(t) for t in transactions],
    )
    return _save_workbook(wb)


def build_transaction_detail_workbook(
    transaction: Transaction,
    result: dict[str, Any],
    specs: dict[str, Any],
    vehicle: dict[str, Any],
    params: dict[str, Any],
) -> io.BytesIO:
    _require_openpyxl()
    audit_buffer = build_audit_workbook(result, specs, vehicle, params)
    wb = openpyxl.load_workbook(audit_buffer)

    ws_info = wb.create_sheet("Passagem", 0)
    _write_key_value_sheet(
        ws_info,
        title=f"Passagem #{transaction.id}",
        subtitle="Metadados da passagem exportada junto com a planilha auditável de cálculo.",
        template_slug=PASSAGEM_SLUG,
        pairs=[
            ("ID", transaction.id),
            ("Placa", transaction.plate),
            ("Contexto", _CONTEXT_LABELS.get(transaction.context, transaction.context)),
            ("UF", transaction.uf),
            ("Digital", "Sim" if transaction.is_digital else "Não"),
            ("Tempo com tag (s)", transaction.elapsed_time_sec),
            ("CO₂ evitado (kg)", transaction.co2_avoided_kg),
            ("Combustível (L)", transaction.fuel_saved_liters),
            ("Tempo economizado (s)", transaction.time_saved_sec),
            ("Economia (R$)", transaction.financial_savings_brl),
            ("Água (L)", transaction.water_saved_liters),
            ("User ID", transaction.user_id),
            ("Veículo ID", transaction.vehicle_id),
            ("Org ID", transaction.organization_id),
            ("Data/Hora", _format_dt(transaction.created_at)),
        ],
    )

    return _save_workbook(wb)
