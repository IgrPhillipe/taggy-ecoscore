"""
Planilha auditável — valores calculados usam fórmulas Excel que referenciam
as premissas. Altere qualquer premissa e todos os resultados se atualizam.
"""

from __future__ import annotations

import io
from typing import Any, Dict

from src.constants.ludic_metaphors import METAPHOR_IDS_ORDER, METAPHOR_LABELS, METAPHOR_SOURCES
from src.constants.workbook_exports import (
    CALCULOS_TEMPLATE_SLUG,
    workbook_export_version_label,
    workbook_template_label,
)

try:
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.worksheet.worksheet import Worksheet
    _HAS_OPENPYXL = True
except ImportError:
    _HAS_OPENPYXL = False

# ── Palette ───────────────────────────────────────────────────────────────────
_C_HEADER_BG = "1A3A4A"
_C_HEADER_FG = "FFFFFF"
_C_WARN_BG   = "FFF3CD"
_C_ALT_BG    = "F5F8FA"
_C_TOTAL_BG  = "D4EDDA"
_C_BORDER    = "CCCCCC"

_FONT_TITLE  = lambda: Font(bold=True, name="Calibri", size=14)
_FONT_SMALL  = lambda: Font(name="Calibri", size=9, color="555555")

_FILL_HEADER = lambda: PatternFill("solid", fgColor=_C_HEADER_BG)
_FILL_WARN   = lambda: PatternFill("solid", fgColor=_C_WARN_BG)
_FILL_ALT    = lambda: PatternFill("solid", fgColor=_C_ALT_BG)
_FILL_TOTAL  = lambda: PatternFill("solid", fgColor=_C_TOTAL_BG)

_THIN = lambda: Border(
    left=Side(style="thin", color=_C_BORDER),
    right=Side(style="thin", color=_C_BORDER),
    top=Side(style="thin", color=_C_BORDER),
    bottom=Side(style="thin", color=_C_BORDER),
)

def _header_row(ws: "Worksheet", row: int, cols: list[str]) -> None:
    for i, v in enumerate(cols, 1):
        c = ws.cell(row=row, column=i, value=v)
        c.font = Font(bold=True, color=_C_HEADER_FG, name="Calibri", size=11)
        c.fill = _FILL_HEADER()
        c.border = _THIN()
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _set_col_widths(ws: "Worksheet", widths: list[int]) -> None:
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _cell(ws: "Worksheet", row: int, col: int, value, *,
          bold=False, fill=None, fmt=None, wrap=False, size=10, color="000000"):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(bold=bold, name="Calibri", size=size, color=color)
    c.border = _THIN()
    if fill:
        c.fill = fill
    if fmt:
        c.number_format = fmt
    c.alignment = Alignment(wrap_text=wrap, vertical="top")
    return c


# ── Source URLs — hyperlinks para a coluna Fonte na aba Premissas ─────────────
_SOURCE_URLS: dict[str, str] = {
    "emission_factors": "https://fgvcli.fgv.br/ghg",
    "gwp100":           "https://www.ipcc.ch/report/ar6/wg1/",
    "blend_factors":    "https://www.planalto.gov.br/ccivil_03/_ato2023-2026/2024/lei/l14993.htm",
    "idle_rates":       "https://blog.edenredmobilidade.com.br/gestao-de-frotas/impacto-da-conducao-no-consumo-de-combustivel/",
    "paper_impact":     "https://ecoinvent.org/database/",
    "fuel_prices":      "https://www.gov.br/anp/pt-br/assuntos/precos-e-defesa-da-concorrencia/precos/levantamento-de-precos",
}

# ── Premises list ─────────────────────────────────────────────────────────────
# (key, label, unit, source_key, warn)
_PREMISES: list[tuple] = [
    ("emission_factor_gasolina_c_base",      "Fator CO₂ gasolina A pura (base)",                  "kg CO₂/L",   "emission_factors", False),
    ("blend_etanol_pct",                      "Blend etanol na gasolina C (E27-E30)",               "%",          "blend_factors",    False),
    ("emission_factor_gasolina_c_comercial",  "Fator CO₂ gasolina C comercial (blendado)",         "kg CO₂/L",   "emission_factors", False),
    ("ch4_factor_gasolina_c",                 "Fator CH4 gasolina C (blendado)",                   "kg CH4/L",   "emission_factors", False),
    ("n2o_factor_gasolina_c",                 "Fator N2O gasolina C (blendado)",                   "kg N2O/L",   "emission_factors", False),
    ("emission_factor_diesel_s10_base",       "Fator CO₂ diesel S10 puro (base)",                  "kg CO₂/L",   "emission_factors", False),
    ("blend_biodiesel_pct",                   "Blend biodiesel no diesel S10 (B14-B15)",            "%",          "blend_factors",    False),
    ("emission_factor_diesel_s10_comercial",  "Fator CO₂ diesel S10 comercial (blendado)",         "kg CO₂/L",   "emission_factors", False),
    ("ch4_factor_diesel_s10",                 "Fator CH4 diesel S10 (blendado)",                   "kg CH4/L",   "emission_factors", False),
    ("n2o_factor_diesel_s10",                 "Fator N2O diesel S10 (blendado)",                   "kg N2O/L",   "emission_factors", False),
    ("emission_factor_etanol",                "Fator CO₂ etanol (biogênico — Escopo 1 = 0)",       "kg CO₂/L",   "emission_factors", False),
    ("ch4_factor_etanol",                     "Fator CH4 etanol",                                  "kg CH4/L",   "emission_factors", False),
    ("n2o_factor_etanol",                     "Fator N2O etanol",                                  "kg N2O/L",   "emission_factors", False),
    ("emission_factor_gnv",                   "Fator CO₂ GNV",                                    "kg CO₂/m³",  "emission_factors", False),
    ("ch4_factor_gnv",                        "Fator CH4 GNV",                                    "kg CH4/m³",  "emission_factors", False),
    ("n2o_factor_gnv",                        "Fator N2O GNV",                                    "kg N2O/m³",  "emission_factors", False),
    ("emission_factor_eletrico_kwh",          "Fator CO₂ rede SIN (veículo elétrico)",             "kg CO₂/kWh", "emission_factors", False),
    ("gwp100_ch4",                            "GWP100 CH4 (IPCC AR6)",                            "—",          "gwp100",           False),
    ("gwp100_n2o",                            "GWP100 N2O (IPCC AR6)",                            "—",          "gwp100",           False),
    ("idle_rate_leve",                        "Taxa consumo idle — leve (L/s)",                    "L/s",        "idle_rates",       True),
    ("idle_rate_pesado",                      "Taxa consumo idle — pesado (L/s)",                  "L/s",        "idle_rates",       True),
    ("idle_rate_gnv",                         "Taxa consumo idle — GNV (m³/s)",                   "m³/s",       "idle_rates",       True),
    ("idle_rate_eletrico",                    "Taxa consumo idle — elétrico (kWh/s)",              "kWh/s",      "idle_rates",       True),
    ("accel_surge_leve",                      "Combustível extra freada+aceleração — leve",        "L",          None,               True),
    ("accel_surge_pesado",                    "Combustível extra freada+aceleração — pesado",      "L",          None,               True),
    ("baseline_pedagio_avg_wait_sec",         "Tempo médio sem tag — pedágio",                    "s",          None,               True),
    ("baseline_estacionamento_avg_wait_sec",  "Tempo médio sem tag — estacionamento",             "s",          None,               True),
    ("paper_co2_per_ticket",                  "CO₂ por ticket de papel (estacionamento)",         "kg CO₂",     "paper_impact",     False),
    ("paper_water_per_ticket",                "Água por ticket de papel (estacionamento)",        "L",          "paper_impact",     False),
    ("fuel_price_brl_per_unit",               "Preço médio do combustível (ANP/UF)",              "R$/unid.",   "fuel_prices",      False),
]

# Precomputed row numbers: first data row = 4 (title=1, note=2, header=3)
_PREM_ROW: dict[str, int] = {key: 4 + i for i, (key, *_) in enumerate(_PREMISES)}


# ── Step row assignments in Sheet 2 ──────────────────────────────────────────
# title=1, vehicle_info=2, header=3, data starts at 4
_SR: dict[str, int] = {
    "baseline":   9,
    "elapsed":    10,
    "time_saved": 11,
    "fuel":       12,
    "co2_fossil": 13,
    "ch4_abs":    14,
    "ch4_co2e":   15,
    "n2o_abs":    16,
    "n2o_co2e":   17,
    "scope1":     18,
    "scope2":     19,
    "paper":      20,
    "total":      21,
}


# ── Fuel helpers ──────────────────────────────────────────────────────────────

def _idle_key(fuel_type: str, category: str) -> str:
    if fuel_type == "eletrico":
        return "idle_rate_eletrico"
    if fuel_type == "gnv":
        return "idle_rate_gnv"
    return "idle_rate_pesado" if category == "pesado" else "idle_rate_leve"


def _accel_key(fuel_type: str, category: str) -> str | None:
    if fuel_type in ("gnv", "eletrico"):
        return None
    return "accel_surge_pesado" if category == "pesado" else "accel_surge_leve"


def _ef_scope1_key(fuel_type: str) -> str | None:
    return {
        "gasolina_c": "emission_factor_gasolina_c_comercial",
        "diesel_s10": "emission_factor_diesel_s10_comercial",
        "gnv":        "emission_factor_gnv",
    }.get(fuel_type)


def _ch4_key(fuel_type: str) -> str | None:
    return {
        "gasolina_c": "ch4_factor_gasolina_c",
        "diesel_s10": "ch4_factor_diesel_s10",
        "etanol":     "ch4_factor_etanol",
        "gnv":        "ch4_factor_gnv",
    }.get(fuel_type)


def _n2o_key(fuel_type: str) -> str | None:
    return {
        "gasolina_c": "n2o_factor_gasolina_c",
        "diesel_s10": "n2o_factor_diesel_s10",
        "etanol":     "n2o_factor_etanol",
        "gnv":        "n2o_factor_gnv",
    }.get(fuel_type)


def _fuel_unit(fuel_type: str) -> str:
    return "kWh" if fuel_type == "eletrico" else ("m³" if fuel_type == "gnv" else "L")


# ── Dynamic IF-chain formula helpers (dropdown-aware) ─────────────────────────

def _dyn_idle_ref(prem_rows: dict, fuel_cell: str, cat_cell: str) -> str:
    return (
        f"IF({fuel_cell}=\"eletrico\",{_pref(prem_rows['idle_rate_eletrico'])},"
        f"IF({fuel_cell}=\"gnv\",{_pref(prem_rows['idle_rate_gnv'])},"
        f"IF({cat_cell}=\"pesado\",{_pref(prem_rows['idle_rate_pesado'])},{_pref(prem_rows['idle_rate_leve'])})))"
    )


def _dyn_accel_ref(prem_rows: dict, fuel_cell: str, cat_cell: str) -> str:
    return (
        f"IF(OR({fuel_cell}=\"gnv\",{fuel_cell}=\"eletrico\"),0,"
        f"IF({cat_cell}=\"pesado\",{_pref(prem_rows['accel_surge_pesado'])},{_pref(prem_rows['accel_surge_leve'])}))"
    )


def _dyn_ef_scope1_ref(prem_rows: dict, fuel_cell: str) -> str:
    return (
        f"IF({fuel_cell}=\"gasolina_c\",{_pref(prem_rows['emission_factor_gasolina_c_comercial'])},"
        f"IF({fuel_cell}=\"diesel_s10\",{_pref(prem_rows['emission_factor_diesel_s10_comercial'])},"
        f"IF({fuel_cell}=\"gnv\",{_pref(prem_rows['emission_factor_gnv'])},0)))"
    )


def _dyn_ch4_ref(prem_rows: dict, fuel_cell: str) -> str:
    return (
        f"IF({fuel_cell}=\"gasolina_c\",{_pref(prem_rows['ch4_factor_gasolina_c'])},"
        f"IF({fuel_cell}=\"diesel_s10\",{_pref(prem_rows['ch4_factor_diesel_s10'])},"
        f"IF({fuel_cell}=\"etanol\",{_pref(prem_rows['ch4_factor_etanol'])},"
        f"IF({fuel_cell}=\"gnv\",{_pref(prem_rows['ch4_factor_gnv'])},0))))"
    )


def _dyn_n2o_ref(prem_rows: dict, fuel_cell: str) -> str:
    return (
        f"IF({fuel_cell}=\"gasolina_c\",{_pref(prem_rows['n2o_factor_gasolina_c'])},"
        f"IF({fuel_cell}=\"diesel_s10\",{_pref(prem_rows['n2o_factor_diesel_s10'])},"
        f"IF({fuel_cell}=\"etanol\",{_pref(prem_rows['n2o_factor_etanol'])},"
        f"IF({fuel_cell}=\"gnv\",{_pref(prem_rows['n2o_factor_gnv'])},0))))"
    )


def _dyn_scope2_ref(prem_rows: dict, fuel_cell: str) -> str:
    return f"IF({fuel_cell}=\"eletrico\",{_pref(prem_rows['emission_factor_eletrico_kwh'])},0)"


def _dyn_baseline_ref(prem_rows: dict, ctx_cell: str) -> str:
    return (
        f"IF({ctx_cell}=\"pedagio\","
        f"{_pref(prem_rows['baseline_pedagio_avg_wait_sec'])},"
        f"{_pref(prem_rows['baseline_estacionamento_avg_wait_sec'])})"
    )


def _dyn_paper_ref(prem_rows: dict, ctx_cell: str) -> str:
    return f"IF({ctx_cell}=\"estacionamento\",{_pref(prem_rows['paper_co2_per_ticket'])},0)"


# ── Cross-sheet formula refs ──────────────────────────────────────────────────
_P  = "'1. Premissas'"
_S2 = "'2. Passo a Passo'"

# Editable param cell refs — intra-Sheet2
_FUEL  = "$D$3"
_CAT   = "$D$4"
_CTX   = "$D$5"
_PRICE = "$D$6"

# Editable param cell refs — cross-sheet (used in Sheets 3 and 4)
_S2_FUEL  = f"{_S2}!$D$3"
_S2_CAT   = f"{_S2}!$D$4"
_S2_CTX   = f"{_S2}!$D$5"
_S2_PRICE = f"{_S2}!$D$6"


def _pref(row: int) -> str:
    return f"{_P}!C{row}"


def _s2e(row: int) -> str:
    return f"{_S2}!E{row}"


def _sens_total(
    time_expr: str,
    idle_expr: str,
    accel_term: str,
    ef_expr: str,
    ch4_expr: str,
    n2o_expr: str,
    gwp_ch4_expr: str,
    gwp_n2o_expr: str,
    scope2_ef_expr: str,
    paper_expr: str,
) -> str:
    """Build Excel formula for total CO₂e, varying one parameter at a time."""
    fuel = f"({time_expr}*{idle_expr}{accel_term})"
    scope1 = f"({fuel}*({ef_expr}+{ch4_expr}*{gwp_ch4_expr}+{n2o_expr}*{gwp_n2o_expr}))"
    scope2 = f"({fuel}*{scope2_ef_expr})" if scope2_ef_expr else "0"
    return f"={scope1}+{scope2}+({paper_expr})"


# ── Sheet 1: Premissas ────────────────────────────────────────────────────────

def _build_premises_sheet(ws: "Worksheet", specs: Dict[str, Any], fuel_price: float = 0.0, fuel_price_meta: dict | None = None) -> dict[str, int]:
    ws.title = "1. Premissas"

    ws.merge_cells("A1:G1")
    ws["A1"].value = "Premissas e Fontes — Valores Brutos (altere aqui para atualizar todos os cálculos)"
    ws["A1"].font = _FONT_TITLE()
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:G2")
    ws["A2"].value = (
        "⚠️  Linhas em amarelo = estimativa sem fonte pública oficial brasileira. "
        "Todas as outras sheets usam fórmulas que referenciam a coluna C desta sheet — "
        "altere qualquer valor e os resultados se atualizam automaticamente."
    )
    ws["A2"].font = Font(name="Calibri", size=9, italic=True, color="7D6608")
    ws["A2"].fill = PatternFill("solid", fgColor="FFF9E6")
    ws["A2"].alignment = Alignment(wrap_text=True)
    ws.row_dimensions[2].height = 32

    _header_row(ws, 3, ["Chave", "Descrição", "Valor", "Unidade", "Fonte", "Ano", "Notas"])

    ef     = specs.get("emission_factors", {})
    ch4    = specs.get("ch4_factors", {})
    n2o    = specs.get("n2o_factors", {})
    gwp    = specs.get("gwp100", {})
    idle   = specs.get("idle_rates", {})
    bases  = specs.get("baselines", {})
    paper  = specs.get("paper_impact", {})
    blend  = specs.get("blend", {})
    accel  = specs.get("accel_surge", {})
    src    = specs.get("sources", {})

    def _src(sk: str | None) -> tuple[str, str]:
        if not sk:
            return ("Premissa declarada — sem dado público disponível", "—")
        if sk == "fuel_prices":
            fmeta = fuel_price_meta or {}
            return (fmeta.get("source", "ANP"), "—")
        return (src.get(sk, "—"), str(src.get(f"{sk}_year", "—")))

    blend_eth = blend.get("etanol_pct", 0.27)
    blend_bio = blend.get("biodiesel_pct", 0.14)

    gas_base = round(ef.get("gasolina_c", 0) / (1 - blend_eth), 6) if blend_eth < 1 else ef.get("gasolina_c", 0)
    die_base = round(ef.get("diesel_s10", 0) / (1 - blend_bio), 6) if blend_bio < 1 else ef.get("diesel_s10", 0)

    notes_map: dict[str, str] = {
        "emission_factor_gasolina_c_base": (
            "Quanto CO₂ é liberado ao queimar 1 litro de gasolina A (pura, sem mistura). "
            "A gasolina que abastecemos no posto (gasolina C) é uma mistura desta base com etanol."
        ),
        "emission_factor_gasolina_c_comercial": (
            f"Fator já com o blend de {round(blend_eth*100,1)}% de etanol (Lei 14.993/2024 — E30). "
            "Como etanol é biogênico, o blend reduz o CO₂ fóssil por litro. Este é o valor usado nos cálculos."
        ),
        "emission_factor_diesel_s10_base": (
            "Emissão de CO₂ de 1 litro de diesel puro (sem biodiesel). "
            "Diesel S10 = baixo teor de enxofre (10 ppm), padrão atual para veículos pesados."
        ),
        "emission_factor_diesel_s10_comercial": (
            f"Fator já com o blend de {round(blend_bio*100,1)}% de biodiesel (B15, Resolução CNPE). "
            "Biodiesel é renovável e reduz o CO₂ fóssil por litro. Este é o valor usado nos cálculos."
        ),
        "emission_factor_etanol": (
            "CO₂ do etanol é classificado como biogênico (a planta absorveu CO₂ ao crescer). "
            "Pelo GHG Protocol, emissões biogênicas não contam no Escopo 1 — por isso o fator de Escopo 1 é zero."
        ),
        "ch4_factor_gasolina_c":  "Metano (CH4) emitido por litro. Apesar de pequeno em massa, CH4 aquece ~27,9× mais que CO₂ (ver GWP100 abaixo).",
        "n2o_factor_gasolina_c":  "Óxido nitroso (N2O) emitido por litro. Em massa mínima, mas 273× mais potente que CO₂ em 100 anos.",
        "ch4_factor_diesel_s10":  "Metano (CH4) emitido por litro de diesel S10.",
        "n2o_factor_diesel_s10":  "Óxido nitroso (N2O) emitido por litro de diesel S10.",
        "emission_factor_etanol": "CO₂ biogênico do etanol — Escopo 1 = 0 pelo GHG Protocol (carbono absorvido na fase de cultivo).",
        "ch4_factor_etanol":      "Metano do etanol — mesmo sendo biogênico, CH4 tem GWP100 alto e entra no CO₂e.",
        "n2o_factor_etanol":      "Óxido nitroso do etanol.",
        "emission_factor_gnv": (
            "GNV = Gás Natural Veicular (metano comprimido). Unidade em m³ (metro cúbico), não litros."
        ),
        "ch4_factor_gnv":  "Metano fugitivo do GNV — além do que é queimado, parte escapa como CH4 puro.",
        "n2o_factor_gnv":  "Óxido nitroso emitido pelo GNV.",
        "emission_factor_eletrico_kwh": (
            "Fator da rede elétrica brasileira (SIN) em kg CO₂ por kWh consumido. "
            "Atualizado anualmente pela FGV/ONS com base na matriz energética. "
            "Conta como Escopo 2 (emissões indiretas da geração de energia)."
        ),
        "gwp100_ch4": (
            "GWP100 = Potencial de Aquecimento Global em 100 anos. "
            "1 kg de CH4 = 27,9 kg de CO₂e. Usado para converter CH4 em CO₂ equivalente. Fonte: IPCC AR6 2021."
        ),
        "gwp100_n2o": (
            "GWP100 do N2O. 1 kg de N2O = 273 kg de CO₂e. "
            "Apesar de emitido em quantidade minúscula, impacta significativamente o total de CO₂e."
        ),
        "idle_rate_leve": (
            "Litros de combustível consumidos por segundo com motor ligado parado (marcha lenta). "
            "Veículos leves = carros de passeio. Fonte: Contele Rastreador (BR) — 1,5 L/h para frota leve brasileira."
        ),
        "idle_rate_pesado": (
            "Mesmo conceito para veículos pesados (caminhões, ônibus). "
            "Fonte: Edenred Mobilidade (BR) — 4 L/h para veículos pesados. Consumo idle é proporcionalmente maior por cilindrada."
        ),
        "idle_rate_gnv": "Taxa de consumo idle para veículos GNV, estimada por conversão energética equivalente.",
        "idle_rate_eletrico": "Consumo elétrico do sistema com veículo parado (ar condicionado, eletrônica). Estimativa — sem fonte oficial.",
        "accel_surge_leve": (
            "Combustível extra gasto no ciclo completo de frenagem + aceleração ao parar manualmente. "
            "Atualmente zerado (conservador) — pode ser atualizado com dados de campo."
        ),
        "accel_surge_pesado": (
            "Mesmo conceito para veículos pesados. A frenagem e retomada de velocidade de um caminhão "
            "consome significativamente mais combustível que em leves."
        ),
        "baseline_pedagio_avg_wait_sec": (
            "Tempo médio estimado de uma passagem SEM tag no pedágio: "
            "fila + parar + pagar + troco + partir. Premissa declarada (~180s = 3 min) — sem dado oficial ANTT/ABCR."
        ),
        "baseline_estacionamento_avg_wait_sec": (
            "Tempo estimado para retirar ticket, aguardar cancela abrir e partir. "
            "Premissa declarada (~120s) — sem dado oficial."
        ),
        "paper_co2_per_ticket": (
            "CO₂ emitido na produção de 1 ticket de papel térmico (estacionamento). "
            "Aplica-se apenas ao contexto 'estacionamento' com tag digital (sem papel)."
        ),
        "paper_water_per_ticket": (
            "Água consumida na produção de 1 ticket de papel (extração de celulose, branqueamento). "
            "Aplica-se apenas a estacionamento digital."
        ),
        "fuel_price_brl_per_unit": (
            f"UF: {(fuel_price_meta or {}).get('uf', '?')} | "
            f"Unidade: {(fuel_price_meta or {}).get('unit', 'L')} | "
            f"Fonte: {(fuel_price_meta or {}).get('source', 'ANP')}. "
            "Preço médio ANP por estado — atualizado automaticamente a cada passagem."
        ),
    }

    values_map: dict[str, Any] = {
        "emission_factor_gasolina_c_base":      gas_base,
        "blend_etanol_pct":                     f"{round(blend_eth * 100, 1)}%",
        "emission_factor_gasolina_c_comercial": round(ef.get("gasolina_c", 0), 6),
        "ch4_factor_gasolina_c":                round(ch4.get("gasolina_c", 0), 8),
        "n2o_factor_gasolina_c":                round(n2o.get("gasolina_c", 0), 8),
        "emission_factor_diesel_s10_base":      die_base,
        "blend_biodiesel_pct":                  f"{round(blend_bio * 100, 1)}%",
        "emission_factor_diesel_s10_comercial": round(ef.get("diesel_s10", 0), 6),
        "ch4_factor_diesel_s10":                round(ch4.get("diesel_s10", 0), 8),
        "n2o_factor_diesel_s10":                round(n2o.get("diesel_s10", 0), 8),
        "emission_factor_etanol":               round(ef.get("etanol", 0), 6),
        "ch4_factor_etanol":                    round(ch4.get("etanol", 0), 8),
        "n2o_factor_etanol":                    round(n2o.get("etanol", 0), 8),
        "emission_factor_gnv":                  round(ef.get("gnv", 0), 6),
        "ch4_factor_gnv":                       round(ch4.get("gnv", 0), 8),
        "n2o_factor_gnv":                       round(n2o.get("gnv", 0), 8),
        "emission_factor_eletrico_kwh":         round(ef.get("eletrico_kwh", 0), 6),
        "gwp100_ch4":                           gwp.get("ch4", 27.9),
        "gwp100_n2o":                           gwp.get("n2o", 273.0),
        "idle_rate_leve":                       idle.get("leve", 0),
        "idle_rate_pesado":                     idle.get("pesado", 0),
        "idle_rate_gnv":                        idle.get("gnv", 0),
        "idle_rate_eletrico":                   idle.get("eletrico", 0),
        "accel_surge_leve":                     round(accel.get("leve", 0), 6),
        "accel_surge_pesado":                   round(accel.get("pesado", 0), 6),
        "baseline_pedagio_avg_wait_sec":        bases.get("pedagio", {}).get("avg_wait_sec", 180),
        "baseline_estacionamento_avg_wait_sec": bases.get("estacionamento", {}).get("avg_wait_sec", 120),
        "paper_co2_per_ticket":                 round(paper.get("co2_per_ticket", 0), 6),
        "paper_water_per_ticket":               round(paper.get("water_per_ticket", 0), 4),
        "fuel_price_brl_per_unit":              round(float(fuel_price or 0), 4),
    }

    prem_rows: dict[str, int] = {}

    for row_i, (key, desc, unit, src_key, warn) in enumerate(_PREMISES, 4):
        prem_rows[key] = row_i
        val = values_map.get(key, 0)
        src_text, src_year = _src(src_key)
        note = notes_map.get(key, "")
        fill = _FILL_WARN() if warn else (_FILL_ALT() if row_i % 2 == 0 else None)

        for col, v in enumerate([key, desc, val, unit, src_text, src_year, note], 1):
            c = ws.cell(row=row_i, column=col, value=v)
            c.font = Font(name="Calibri", size=10, bold=(col == 1))
            c.border = _THIN()
            c.alignment = Alignment(wrap_text=True, vertical="top")
            if fill:
                c.fill = fill
            if col == 3 and isinstance(val, float):
                c.number_format = "0.00######"
            # Hyperlink na coluna Fonte (col 5)
            if col == 5 and src_key and src_key in _SOURCE_URLS:
                c.hyperlink = _SOURCE_URLS[src_key]
                c.font = Font(name="Calibri", size=10, color="0563C1", underline="single")
                if fill:
                    c.fill = fill

    _set_col_widths(ws, [38, 50, 12, 12, 52, 8, 60])
    ws.freeze_panes = "A4"

    # ── Seção de Métricas Lúdicas ────────────────────────────────────────────
    ludic_start = 4 + len(_PREMISES) + 2
    ws.merge_cells(f"A{ludic_start}:G{ludic_start}")
    c_h = ws.cell(row=ludic_start, column=1, value="Métricas Lúdicas — Conversões e Fontes")
    c_h.font = Font(bold=True, name="Calibri", size=12, color=_C_HEADER_FG)
    c_h.fill = _FILL_HEADER()
    c_h.border = _THIN()
    c_h.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[ludic_start].height = 22

    ws.merge_cells(f"A{ludic_start+1}:G{ludic_start+1}")
    c_note = ws.cell(
        row=ludic_start + 1, column=1,
        value="Conversões usadas para traduzir impacto técnico em linguagem cotidiana. "
              "Valores de conversão vêm do banco de dados (configuráveis). Fontes abaixo.",
    )
    c_note.font = Font(name="Calibri", size=9, italic=True, color="555555")
    c_note.alignment = Alignment(wrap_text=True)
    ws.row_dimensions[ludic_start + 1].height = 20

    _header_row(ws, ludic_start + 2, ["Eixo", "Metáfora", "Unidade de Conversão", "Valor (BD)", "Fonte", "URL"])

    ludic_units = specs.get("ludic_metaphor_units") or {}
    row_l = ludic_start + 3
    for axis, ids in METAPHOR_IDS_ORDER.items():
        axis_labels = METAPHOR_LABELS.get(axis, {})
        axis_sources = METAPHOR_SOURCES.get(axis, {})
        for mid in ids:
            label = axis_labels.get(mid, mid)
            val = (ludic_units.get(axis) or {}).get(mid) or (
                {"carbon": {"tree_year": 15.0, "burger": 2.5, "km_car": 0.12},
                 "water": {"shower_8min": 60.0, "drinking_day": 2.0, "flush": 6.0},
                 "paper": {"ream_a4": 500.0, "notebook": 50.0, "toilet_roll": 150.0}
                 }.get(axis, {}).get(mid, "—")
            )
            src_info = axis_sources.get(mid, ("—", None))
            src_text, src_url = src_info if isinstance(src_info, tuple) else (src_info, None)
            fill = _FILL_ALT() if row_l % 2 == 0 else None
            for col, v in enumerate([axis, label, "por unidade da metáfora", val, src_text, src_url or ""], 1):
                c = ws.cell(row=row_l, column=col, value=v)
                c.font = Font(name="Calibri", size=10)
                c.border = _THIN()
                c.alignment = Alignment(wrap_text=True, vertical="top")
                if fill:
                    c.fill = fill
                if col == 6 and src_url:
                    c.hyperlink = src_url
                    c.font = Font(name="Calibri", size=10, color="0563C1", underline="single")
            row_l += 1

    return prem_rows


# ── Sheet 2: Passo a Passo ────────────────────────────────────────────────────

def _build_steps_sheet(
    ws: "Worksheet",
    vehicle: Dict[str, Any],
    params: Dict[str, Any],
    prem_rows: dict[str, int],
) -> None:
    ws.title = "2. Passo a Passo"

    fuel_type  = vehicle.get("fuel_type", "gasolina_c")
    category   = vehicle.get("category", "leve")
    context    = params.get("context", "pedagio")
    elapsed    = int(params.get("elapsed_time", 30))

    funit = _fuel_unit(fuel_type)
    r_gch4 = prem_rows["gwp100_ch4"]
    r_gn2o = prem_rows["gwp100_n2o"]

    R = _SR

    # Dynamic IF-chain refs (intra-sheet dropdowns _FUEL/$D$3, _CAT/$D$4, _CTX/$D$5)
    idle_ref  = _dyn_idle_ref(prem_rows, _FUEL, _CAT)
    accel_ref = _dyn_accel_ref(prem_rows, _FUEL, _CAT)
    ef_ref    = _dyn_ef_scope1_ref(prem_rows, _FUEL)
    ch4_ref   = _dyn_ch4_ref(prem_rows, _FUEL)
    n2o_ref   = _dyn_n2o_ref(prem_rows, _FUEL)
    bl_ref    = _dyn_baseline_ref(prem_rows, _CTX)
    paper_ref = _dyn_paper_ref(prem_rows, _CTX)

    # (key, description, formula_desc, result_value_or_formula, unit)
    steps = [
        ("baseline_sem_tag",
         "Tempo médio sem usar tag",
         f"IF(Contexto=pedagio, Premissas linha {prem_rows['baseline_pedagio_avg_wait_sec']}, linha {prem_rows['baseline_estacionamento_avg_wait_sec']})",
         f"={bl_ref}",
         "s"),

        ("tempo_com_tag",
         "Tempo real da passagem com tag (input do sistema)",
         "Parâmetro da requisição — medido",
         elapsed,
         "s"),

        ("tempo_salvo",
         "Tempo economizado",
         f"MAX(0, E{R['baseline']} − E{R['elapsed']})",
         f"=MAX(0,E{R['baseline']}-E{R['elapsed']})",
         "s"),

        ("combustivel_evitado",
         f"Combustível não consumido (L/m³/kWh conforme combustível selecionado)",
         f"E{R['time_saved']} × idle_rate(combustível, categoria) + accel_surge(combustível, categoria)",
         f"=E{R['time_saved']}*{idle_ref}+{accel_ref}",
         funit),

        ("co2_fossil",
         "CO₂ fóssil evitado (0 para etanol biogênico e elétrico)",
         f"E{R['fuel']} × ef_scope1(combustível)",
         f"=E{R['fuel']}*{ef_ref}",
         "kg CO₂"),

        ("ch4_absoluto",
         "CH4 evitado (massa) — 0 para elétrico",
         f"E{R['fuel']} × ch4_factor(combustível)",
         f"=E{R['fuel']}*{ch4_ref}",
         "kg CH4"),

        ("ch4_co2e",
         "CH4 em CO₂e",
         f"E{R['ch4_abs']} × Premissas!C{r_gch4} (GWP100)",
         f"=E{R['ch4_abs']}*{_pref(r_gch4)}",
         "kg CO₂e"),

        ("n2o_absoluto",
         "N2O evitado (massa) — 0 para elétrico",
         f"E{R['fuel']} × n2o_factor(combustível)",
         f"=E{R['fuel']}*{n2o_ref}",
         "kg N2O"),

        ("n2o_co2e",
         "N2O em CO₂e",
         f"E{R['n2o_abs']} × Premissas!C{r_gn2o} (GWP100)",
         f"=E{R['n2o_abs']}*{_pref(r_gn2o)}",
         "kg CO₂e"),

        ("co2e_scope1",
         "CO₂e Escopo 1 (combustão direta)",
         f"E{R['co2_fossil']} + E{R['ch4_co2e']} + E{R['n2o_co2e']}",
         f"=E{R['co2_fossil']}+E{R['ch4_co2e']}+E{R['n2o_co2e']}",
         "kg CO₂e"),

        ("co2e_scope2",
         "CO₂e Escopo 2 (rede elétrica — só EV, 0 para outros)",
         f"E{R['fuel']} × ef_eletrico se combustível=eletrico, senão 0",
         f"=E{R['fuel']}*{_dyn_scope2_ref(prem_rows, _FUEL)}",
         "kg CO₂e"),

        ("paper_co2_avoided",
         "CO₂ ticket de papel evitado (só estacionamento)",
         f"IF(contexto=estacionamento, Premissas linha {prem_rows['paper_co2_per_ticket']}, 0)",
         f"={paper_ref}",
         "kg CO₂"),

        ("TOTAL_EVITADO",
         "TOTAL CO₂e EVITADO",
         f"E{R['scope1']} + E{R['scope2']} + E{R['paper']}",
         f"=E{R['scope1']}+E{R['scope2']}+E{R['paper']}",
         "kg CO₂e"),
    ]

    # ── Row 1: title ─────────────────────────────────────────────────────────
    ws.merge_cells("A1:F1")
    ws["A1"].value = "Passo a Passo — Emissões Evitadas por Passagem com Tag"
    ws["A1"].font = _FONT_TITLE()
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 28

    # ── Rows 2-5: Parâmetros editáveis ───────────────────────────────────────
    ws.merge_cells("A2:F2")
    ws["A2"].value = "⚙️  Parâmetros de Simulação — altere os valores abaixo e todas as fórmulas recalculam"
    ws["A2"].font = Font(name="Calibri", size=10, bold=True, color="7D6608")
    ws["A2"].fill = PatternFill("solid", fgColor="FFF9E6")
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 22

    fuel_price = float(params.get("fuel_price_brl_per_unit") or 0)

    dropdown_params = [
        ("Combustível", fuel_type,
         '"gasolina_c,diesel_s10,etanol,gnv,eletrico"',
         "Selecione o tipo de combustível",
         "Combustível"),
        ("Categoria", category,
         '"leve,pesado"',
         "Selecione a categoria do veículo",
         "Categoria"),
        ("Contexto", context,
         '"pedagio,estacionamento"',
         "Selecione o contexto da passagem",
         "Contexto"),
    ]

    for i, (label, default_val, dv_formula, dv_prompt, dv_title) in enumerate(dropdown_params, 3):
        c_label = ws.cell(row=i, column=1, value=label)
        c_label.font = Font(name="Calibri", size=10, bold=True)
        c_label.border = _THIN()
        c_label.fill = _FILL_ALT()

        ws.merge_cells(f"B{i}:C{i}")
        c_hint = ws.cell(row=i, column=2,
                         value="← selecione no dropdown (seta na célula D)")
        c_hint.font = Font(name="Calibri", size=9, italic=True, color="888888")
        c_hint.border = _THIN()
        c_hint.fill = _FILL_ALT()

        c_val = ws.cell(row=i, column=4, value=default_val)
        c_val.font = Font(name="Calibri", size=10, bold=True, color="1A3A4A")
        c_val.border = _THIN()
        c_val.fill = PatternFill("solid", fgColor="E8F4FD")
        c_val.alignment = Alignment(horizontal="center")

        dv = DataValidation(
            type="list",
            formula1=dv_formula,
            allow_blank=False,
            showDropDown=False,
        )
        dv.prompt = dv_prompt
        dv.promptTitle = dv_title
        ws.add_data_validation(dv)
        dv.add(ws.cell(row=i, column=4))

        ws.row_dimensions[i].height = 20

    # ── Row 6: preço combustível (campo numérico livre) ───────────────────────
    c_plabel = ws.cell(row=6, column=1, value="Preço Combustível (R$/unid.)")
    c_plabel.font = Font(name="Calibri", size=10, bold=True)
    c_plabel.border = _THIN()
    c_plabel.fill = _FILL_ALT()

    ws.merge_cells("B6:C6")
    c_phint = ws.cell(row=6, column=2, value="← edite o valor numérico (R$ por L / m³ / kWh)")
    c_phint.font = Font(name="Calibri", size=9, italic=True, color="888888")
    c_phint.border = _THIN()
    c_phint.fill = _FILL_ALT()

    c_pval = ws.cell(row=6, column=4, value=round(fuel_price, 4))
    c_pval.font = Font(name="Calibri", size=10, bold=True, color="1A3A4A")
    c_pval.border = _THIN()
    c_pval.fill = PatternFill("solid", fgColor="E8F4FD")
    c_pval.alignment = Alignment(horizontal="center")
    c_pval.number_format = "R$ #,##0.0000"
    ws.row_dimensions[6].height = 20

    # ── Row 7: vehicle info ───────────────────────────────────────────────────
    ws.merge_cells("A7:F7")
    ws["A7"].value = (
        f"Veículo: {vehicle.get('model') or 'N/A'} | UF: {params.get('uf','?')} | "
        f"Tempo passagem: {elapsed}s  |  "
        f"(valores padrão acima refletem os dados do veículo — editáveis para simulação)"
    )
    ws["A7"].font = _FONT_SMALL()
    ws.row_dimensions[7].height = 20

    # ── Row 8: column header ──────────────────────────────────────────────────
    _header_row(ws, 8, ["Passo", "Descrição", "Fórmula / Referência", "Referências (células)", "Resultado", "Unidade"])

    # ── Rows 9+: steps ────────────────────────────────────────────────────────
    for row_i, (key, desc, formula_desc, result_val, unit) in enumerate(steps, 9):
        is_total = key == "TOTAL_EVITADO"
        fill = _FILL_TOTAL() if is_total else (_FILL_ALT() if row_i % 2 == 0 else None)

        for col, v in enumerate([key, desc, formula_desc, "", result_val, unit], 1):
            c = ws.cell(row=row_i, column=col, value=v)
            c.font = Font(name="Calibri", size=10, bold=is_total)
            c.border = _THIN()
            c.alignment = Alignment(wrap_text=True, vertical="top")
            if fill:
                c.fill = fill
            if col == 5:
                if isinstance(v, str) and v.startswith("="):
                    c.number_format = "0.000000"
                elif isinstance(v, int) and not isinstance(v, bool):
                    c.number_format = "0"

    _set_col_widths(ws, [26, 42, 42, 1, 14, 12])
    ws.freeze_panes = "A9"


# ── Sheet 3: Comparação ───────────────────────────────────────────────────────

def _build_comparison_sheet(
    ws: "Worksheet",
    result: Dict[str, Any],
    vehicle: Dict[str, Any],
    params: Dict[str, Any],
    prem_rows: dict[str, int],
) -> None:
    ws.title = "3. Comparação"

    comp    = result.get("comparison", {})
    without = comp.get("without_tag", {})
    with_t  = comp.get("with_tag", {})
    delta   = comp.get("delta", {})

    r_gch4 = prem_rows["gwp100_ch4"]
    r_gn2o = prem_rows["gwp100_n2o"]
    r_pw   = prem_rows["paper_water_per_ticket"]

    R = _SR

    # Dynamic refs — follow Sheet 2 dropdowns
    idle_ref   = _dyn_idle_ref(prem_rows, _S2_FUEL, _S2_CAT)
    accel_ref  = _dyn_accel_ref(prem_rows, _S2_FUEL, _S2_CAT)
    ef_ref     = _dyn_ef_scope1_ref(prem_rows, _S2_FUEL)
    ch4_ref    = _dyn_ch4_ref(prem_rows, _S2_FUEL)
    n2o_ref    = _dyn_n2o_ref(prem_rows, _S2_FUEL)
    scope2_ref = _dyn_scope2_ref(prem_rows, _S2_FUEL)
    bl_ref     = _dyn_baseline_ref(prem_rows, _S2_CTX)

    ef_chain = f"({ef_ref}+{ch4_ref}*{_pref(r_gch4)}+{n2o_ref}*{_pref(r_gn2o)})"

    fuel_sem  = f"={bl_ref}*{idle_ref}+{accel_ref}"
    fuel_com  = f"={_s2e(R['elapsed'])}*{idle_ref}"

    co2e1_sem = f"=B5*{ef_chain}"
    co2e1_com = f"=C5*{ef_chain}"
    co2e2_sem = f"=B5*{scope2_ref}"
    co2e2_com = f"=C5*{scope2_ref}"

    water_sem = f"=IF({_S2_CTX}=\"estacionamento\",{_pref(r_pw)},0)"
    water_com = "=0"

    ws.merge_cells("A1:D1")
    ws["A1"].value = "Cenário Sem Tag vs. Com Tag — Comparação"
    ws["A1"].font = _FONT_TITLE()
    ws["A1"].alignment = Alignment(horizontal="left")
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:D2")
    ws["A2"].value = (
        "ℹ️  Sem Tag e Com Tag calculados por fórmulas referenciando '1. Premissas' e '2. Passo a Passo'. "
        "Coluna 'Evitado' = Sem Tag − Com Tag."
    )
    ws["A2"].font = Font(name="Calibri", size=9, italic=True, color="1A3A4A")
    ws["A2"].fill = PatternFill("solid", fgColor="E8F4FD")
    ws["A2"].alignment = Alignment(wrap_text=True)
    ws.row_dimensions[2].height = 28

    _header_row(ws, 3, ["Métrica", "Sem Tag", "Com Tag", "Evitado"])

    # (label, sem_val, com_val, delta_raw_or_None)
    # delta_raw=None → use =B{row}-C{row}
    rows: list[tuple] = [
        ("Tempo na passagem (s)",
         f"={bl_ref}",
         f"={_s2e(R['elapsed'])}",
         None),

        ("Combustível (L/m³/kWh conforme combustível)",
         fuel_sem,
         fuel_com,
         None),

        ("CO₂e Escopo 1 (kg)",
         co2e1_sem,
         co2e1_com,
         None),

        ("CO₂e Escopo 2 — só EV (kg)",
         co2e2_sem,
         co2e2_com,
         None),

        ("CO₂ biogênico (kg)",
         without.get("co2_biogenic_kg", 0),
         with_t.get("co2_biogenic_kg", 0),
         delta.get("co2_biogenic_kg", 0)),

        ("Água (L) — ticket de papel",
         water_sem,
         water_com,
         None),

        ("Custo estimado (R$) — combustível",
         f"=B5*{_S2_PRICE}",
         f"=C5*{_S2_PRICE}",
         None),
    ]

    for row_i, (label, sem_val, com_val, delta_raw) in enumerate(rows, 4):
        fill = _FILL_ALT() if row_i % 2 == 0 else None
        d_val = delta_raw if delta_raw is not None else f"=B{row_i}-C{row_i}"

        for col, v in enumerate([label, sem_val, com_val, d_val], 1):
            c = ws.cell(row=row_i, column=col, value=v)
            c.font = Font(name="Calibri", size=10, bold=(col == 1))
            c.border = _THIN()
            if fill:
                c.fill = fill
            if col > 1:
                if isinstance(v, str) and v.startswith("="):
                    c.number_format = "0.0000"
                elif isinstance(v, float):
                    c.number_format = "0.0000"

    _set_col_widths(ws, [34, 18, 18, 18])


# ── Sheet 4: Sensibilidade ────────────────────────────────────────────────────

def _build_sensitivity_sheet(
    ws: "Worksheet",
    vehicle: Dict[str, Any],
    params: Dict[str, Any],
    prem_rows: dict[str, int],
) -> None:
    ws.title = "4. Sensibilidade"

    r_gch4  = prem_rows["gwp100_ch4"]
    r_gn2o  = prem_rows["gwp100_n2o"]

    R = _SR
    elapsed_ref = _s2e(R["elapsed"])

    # Dynamic refs — follow Sheet 2 dropdowns
    idle_ref   = _dyn_idle_ref(prem_rows, _S2_FUEL, _S2_CAT)
    accel_term = f"+{_dyn_accel_ref(prem_rows, _S2_FUEL, _S2_CAT)}"
    ef_ref     = _dyn_ef_scope1_ref(prem_rows, _S2_FUEL)
    ch4_ref    = _dyn_ch4_ref(prem_rows, _S2_FUEL)
    n2o_ref    = _dyn_n2o_ref(prem_rows, _S2_FUEL)
    scope2_ef  = _dyn_scope2_ref(prem_rows, _S2_FUEL)
    paper_expr = _dyn_paper_ref(prem_rows, _S2_CTX)
    bl_ref     = _dyn_baseline_ref(prem_rows, _S2_CTX)

    time_base = f"MAX(0,{bl_ref}-{elapsed_ref})"

    def _vary(param: str, mult: float) -> str:
        if param == "idle":
            return _sens_total(
                time_base, f"({idle_ref})*{mult}", accel_term,
                ef_ref, ch4_ref, n2o_ref,
                _pref(r_gch4), _pref(r_gn2o), scope2_ef, paper_expr)
        if param == "baseline":
            return _sens_total(
                f"MAX(0,({bl_ref})*{mult}-{elapsed_ref})", idle_ref, accel_term,
                ef_ref, ch4_ref, n2o_ref,
                _pref(r_gch4), _pref(r_gn2o), scope2_ef, paper_expr)
        if param == "ef":
            return _sens_total(
                time_base, idle_ref, accel_term,
                f"({ef_ref})*{mult}", ch4_ref, n2o_ref,
                _pref(r_gch4), _pref(r_gn2o), scope2_ef, paper_expr)
        if param == "gwp_ch4":
            return _sens_total(
                time_base, idle_ref, accel_term,
                ef_ref, ch4_ref, n2o_ref,
                f"{_pref(r_gch4)}*{mult}", _pref(r_gn2o), scope2_ef, paper_expr)
        if param == "gwp_n2o":
            return _sens_total(
                time_base, idle_ref, accel_term,
                ef_ref, ch4_ref, n2o_ref,
                _pref(r_gch4), f"{_pref(r_gn2o)}*{mult}", scope2_ef, paper_expr)
        return f"={_s2e(R['total'])}"

    sens_params: list[tuple] = [
        ("idle",     "idle_rate — Taxa idle motor ligado (varia com combustível/categoria)",  True),
        ("baseline", "baseline_wait_sec — Tempo médio sem tag (varia com contexto)",          True),
        ("ef",       "ef_scope1 — Fator CO₂ (0 para etanol/elétrico; varia com combustível)", False),
        ("gwp_ch4",  "gwp100_ch4 — GWP100 CH4 (IPCC AR6)",                                   False),
        ("gwp_n2o",  "gwp100_n2o — GWP100 N2O (IPCC AR6)",                                   False),
    ]

    ws.merge_cells("A1:F1")
    ws["A1"].value = "Análise de Sensibilidade — Variação ±20% e ±50% nos Parâmetros Principais"
    ws["A1"].font = _FONT_TITLE()
    ws["A1"].alignment = Alignment(horizontal="left")
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:F2")
    ws["A2"].value = (
        f"CO₂e base = {_S2}!E{R['total']} kg | "
        "Cada linha recalcula o total variando apenas aquele parâmetro. "
        "⚠️ = premissa sem fonte pública — maior incerteza."
    )
    ws["A2"].font = _FONT_SMALL()

    _header_row(ws, 3, ["Parâmetro", "Base (kg CO₂e)", "−50%", "−20%", "+20%", "+50%"])
    _header_row(ws, 4, ["",          "(atual)",         "(kg CO₂e)", "(kg CO₂e)", "(kg CO₂e)", "(kg CO₂e)"])

    for row_i, (param_key, label, warn) in enumerate(sens_params, 5):
        fill = _FILL_WARN() if warn else (_FILL_ALT() if row_i % 2 == 0 else None)
        label_str = ("⚠️ " if warn else "") + label

        row_vals = [
            label_str,
            f"={_s2e(R['total'])}",
            _vary(param_key, 0.5),
            _vary(param_key, 0.8),
            _vary(param_key, 1.2),
            _vary(param_key, 1.5),
        ]

        for col, v in enumerate(row_vals, 1):
            c = ws.cell(row=row_i, column=col, value=v)
            c.font = Font(name="Calibri", size=10, bold=(col == 1))
            c.border = _THIN()
            if fill:
                c.fill = fill
            if col > 1 and isinstance(v, str) and v.startswith("="):
                c.number_format = "0.0000"

    note_row = 5 + len(sens_params) + 1
    ws.merge_cells(f"A{note_row}:F{note_row}")
    ws[f"A{note_row}"].value = (
        "Nota: baseline_wait_sec é o parâmetro de maior sensibilidade e não tem fonte pública oficial brasileira. "
        "Recomenda-se medição de campo para inventários GHG formais."
    )
    ws[f"A{note_row}"].font = Font(name="Calibri", size=9, italic=True, color="7D6608")
    ws[f"A{note_row}"].fill = PatternFill("solid", fgColor="FFF9E6")
    ws[f"A{note_row}"].alignment = Alignment(wrap_text=True)
    ws.row_dimensions[note_row].height = 40

    _set_col_widths(ws, [50, 14, 14, 14, 14, 14])


# ── Sheet 5: Escala ───────────────────────────────────────────────────────────

def _build_scale_sheet(ws: "Worksheet", fleet_size: int = 1) -> None:
    ws.title = "5. Escala"

    R = _SR

    ws.merge_cells("A1:C1")
    ws["A1"].value = "Projeção de Escala — Emissões Evitadas por Período"
    ws["A1"].font = _FONT_TITLE()
    ws["A1"].alignment = Alignment(horizontal="left")
    ws.row_dimensions[1].height = 28

    # Input cells (raw, editable)
    inputs = [
        ("Frota (veículos):",          fleet_size, "0"),
        ("Passagens/dia por veículo:", 2,           "0"),
        ("Dias úteis/mês:",            20,          "0"),
        ("CO₂e por passagem (kg):",    f"={_s2e(R['total'])}", "0.000000"),
    ]

    for row_i, (label, val, fmt) in enumerate(inputs, 2):
        c_label = ws.cell(row=row_i, column=1, value=label)
        c_label.font = Font(name="Calibri", size=10, bold=True)
        c_label.border = _THIN()

        c_val = ws.cell(row=row_i, column=2, value=val)
        c_val.font = Font(name="Calibri", size=10)
        c_val.border = _THIN()
        c_val.number_format = fmt

    # B2=fleet, B3=passages/day, B4=work_days, B5=co2e_per_passage
    # Row 6: spacer / header
    ws.row_dimensions[6].height = 6

    _header_row(ws, 7, ["Escala", "Passagens", "CO₂e Evitado (kg)", "CO₂e Evitado (ton)"])

    scenarios = [
        ("Por passagem",
         "=1",
         f"=$B$5",
         f"=$B$5/1000"),

        ("Diário (frota × passagens/dia)",
         "=$B$2*$B$3",
         f"=$B$2*$B$3*$B$5",
         f"=$B$2*$B$3*$B$5/1000"),

        ("Mensal (× dias úteis/mês)",
         "=$B$2*$B$3*$B$4",
         f"=$B$2*$B$3*$B$4*$B$5",
         f"=$B$2*$B$3*$B$4*$B$5/1000"),

        ("Anual (× 12 meses)",
         "=$B$2*$B$3*$B$4*12",
         f"=$B$2*$B$3*$B$4*12*$B$5",
         f"=$B$2*$B$3*$B$4*12*$B$5/1000"),

        ("Nacional/dia (est. 5M passagens)",
         "=5000000",
         "=5000000*$B$5",
         "=5000000*$B$5/1000"),

        ("Nacional/ano",
         "=5000000*365",
         "=5000000*365*$B$5",
         "=5000000*365*$B$5/1000"),
    ]

    for row_i, (label, passagens, co2e_kg, co2e_ton) in enumerate(scenarios, 8):
        is_nat = "Nacional" in label
        fill = _FILL_TOTAL() if is_nat else (_FILL_ALT() if row_i % 2 == 0 else None)

        for col, v in enumerate([label, passagens, co2e_kg, co2e_ton], 1):
            c = ws.cell(row=row_i, column=col, value=v)
            c.font = Font(name="Calibri", size=10, bold=is_nat)
            c.border = _THIN()
            if fill:
                c.fill = fill
            if col > 1 and isinstance(v, str) and v.startswith("="):
                c.number_format = "#,##0.000000" if col in (3,) else "#,##0.0000" if col == 4 else "#,##0"

    note_row = 8 + len(scenarios) + 1
    ws.merge_cells(f"A{note_row}:D{note_row}")
    ws[f"A{note_row}"].value = (
        "Nota: volume nacional baseado em estimativa do fluxo diário de pedágios no Brasil. "
        "Dado real via ANTT/ABCR mediante solicitação. "
        "Altere B2-B4 para simular diferentes cenários de frota."
    )
    ws[f"A{note_row}"].font = Font(name="Calibri", size=9, italic=True)
    ws.row_dimensions[note_row].height = 32

    _set_col_widths(ws, [40, 20, 24, 20])


# ── Sheet 0: Glossário ───────────────────────────────────────────────────────

_GLOSSARY: list[tuple[str, str, str]] = [
    # (Termo, Significado simples, Por que importa nesta planilha)
    (
        "CO₂e (CO₂ equivalente)",
        "Unidade que converte todos os gases de efeito estufa em uma escala comum, "
        "usando o CO₂ como referência. Assim, CH4 e N2O podem ser somados ao CO₂.",
        "Todos os resultados de emissão nesta planilha estão em kg CO₂e — "
        "o número já considera CH4 e N2O além do CO₂ puro.",
    ),
    (
        "Escopo 1 (Scope 1)",
        "Emissões diretas — o que sai diretamente do motor do veículo ao queimar combustível. "
        "Combustão de gasolina, diesel, etanol ou GNV.",
        "A maior parte do CO₂e calculado aqui é Escopo 1. "
        "Veículos elétricos têm Escopo 1 = 0.",
    ),
    (
        "Escopo 2 (Scope 2)",
        "Emissões indiretas da eletricidade consumida. Para veículos elétricos, "
        "a usina que gerou o kWh emitiu CO₂ — isso é Escopo 2.",
        "Aplica-se apenas a veículos elétricos nesta planilha. "
        "Fator da rede SIN (Brasil) varia com a matriz energética anual.",
    ),
    (
        "GWP100 (Global Warming Potential)",
        "Potencial de Aquecimento Global em 100 anos. Compara o poder de aquecimento "
        "de um gás em relação ao CO₂ no período de 100 anos.",
        "Usado para converter kg CH4 → kg CO₂e (×27,9) e kg N2O → kg CO₂e (×273). "
        "Fonte: IPCC AR6 2021.",
    ),
    (
        "Fator de Emissão",
        "Quantidade de CO₂ (ou outro gás) emitida por unidade de combustível queimado. "
        "Ex: 2,212 kg CO₂ por litro de gasolina C.",
        "É o multiplicador principal do cálculo. Varia por tipo de combustível e blend.",
    ),
    (
        "Blend / Mistura",
        "Proporção de biocombustível misturado ao combustível fóssil. "
        "Gasolina C = gasolina A + etanol (E30 = 30% etanol). "
        "Diesel S10 = diesel + biodiesel (B15 = 15% biodiesel).",
        "O blend reduz o fator de emissão fóssil. "
        "A proporção é definida por lei federal (ANP/CNPE) e muda periodicamente.",
    ),
    (
        "Idle Rate (Taxa de Marcha Lenta)",
        "Consumo de combustível por segundo com o motor ligado e o veículo parado. "
        "É o quanto o motor consome enquanto está na fila ou aguardando a cancela.",
        "Parâmetro central para calcular o combustível economizado ao usar tag. "
        "Baseado em referências brasileiras do setor (Edenred Mobilidade / Contele Rastreador).",
    ),
    (
        "Baseline (Tempo médio sem tag)",
        "Estimativa do tempo médio que um veículo leva para passar num pedágio ou "
        "estacionamento SEM usar tag (parar, pagar, receber troco, partir).",
        "A diferença entre baseline e o tempo real com tag define o tempo economizado, "
        "que por sua vez determina o combustível e CO₂e evitados. "
        "⚠️ Premissa declarada — sem dado oficial ANTT/ABCR.",
    ),
    (
        "CO₂ Biogênico",
        "CO₂ de origem biológica (plantas, biomassa). O etanol libera CO₂, mas a cana "
        "absorveu esse mesmo CO₂ ao crescer — ciclo neutro pelo GHG Protocol.",
        "CO₂ biogênico não entra no Escopo 1 desta planilha. "
        "Aparece separado para transparência.",
    ),
    (
        "Ecoinvent",
        "Banco de dados global de inventários de ciclo de vida (LCA). "
        "Contém dados de impacto ambiental de materiais e processos industriais.",
        "Fonte do fator de CO₂ e água por ticket de papel térmico (estacionamento).",
    ),
    (
        "ANP (Agência Nacional do Petróleo)",
        "Agência reguladora brasileira que publica semanalmente os preços médios de "
        "combustíveis por estado (UF).",
        "Preço do combustível usado no cálculo de economia financeira vem da ANP.",
    ),
    (
        "SIN (Sistema Interligado Nacional)",
        "Rede elétrica integrada do Brasil, composta por usinas hidrelétricas, "
        "térmicas, eólicas e solares.",
        "O fator de emissão elétrica (Escopo 2) representa a média do SIN, "
        "calculada anualmente pela FGV/ONS.",
    ),
]


def _build_glossary_sheet(ws: "Worksheet") -> None:
    ws.title = "0. Glossário"

    ws.merge_cells("A1:C1")
    ws["A1"].value = "Glossário — Entendendo os Termos desta Planilha"
    ws["A1"].font = _FONT_TITLE()
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 32

    ws.merge_cells("A2:C2")
    ws["A2"].value = (
        "Esta aba explica os conceitos usados nas demais sheets em linguagem acessível. "
        "Não é necessário conhecer contabilidade de carbono para usar esta calculadora."
    )
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="555555")
    ws["A2"].alignment = Alignment(wrap_text=True)
    ws.row_dimensions[2].height = 28

    ws.merge_cells("A3:C3")
    ws["A3"].value = (
        f"Template: {workbook_template_label(CALCULOS_TEMPLATE_SLUG)} · Versão: {workbook_export_version_label()}"
    )
    ws["A3"].font = Font(name="Calibri", size=9, color="555555")
    ws["A3"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[3].height = 18

    _header_row(ws, 4, ["Termo", "O que significa (linguagem simples)", "Por que aparece nesta planilha"])

    for row_i, (term, meaning, relevance) in enumerate(_GLOSSARY, 5):
        fill = _FILL_ALT() if row_i % 2 == 0 else None
        for col, v in enumerate([term, meaning, relevance], 1):
            c = ws.cell(row=row_i, column=col, value=v)
            c.font = Font(name="Calibri", size=10, bold=(col == 1))
            c.border = _THIN()
            c.alignment = Alignment(wrap_text=True, vertical="top")
            if fill:
                c.fill = fill
        ws.row_dimensions[row_i].height = 52

    _set_col_widths(ws, [28, 68, 52])
    ws.freeze_panes = "A5"


# ── Public API ────────────────────────────────────────────────────────────────

def build_audit_workbook(
    result: Dict[str, Any],
    specs: Dict[str, Any],
    vehicle: Dict[str, Any],
    params: Dict[str, Any],
    fleet_size: int = 1,
) -> "io.BytesIO":
    if not _HAS_OPENPYXL:
        raise RuntimeError("openpyxl não está instalado. Execute: uv add openpyxl")

    wb = openpyxl.Workbook()

    fuel_price = float(params.get("fuel_price_brl_per_unit") or 0)
    fuel_price_meta = {
        "source": params.get("fuel_price_source", "ANP"),
        "uf": params.get("fuel_price_uf", params.get("uf", "?")),
        "unit": params.get("fuel_price_unit", "L"),
    }

    # Glossário é a primeira aba (wb.active é criada vazia por padrão)
    ws0 = wb.active
    _build_glossary_sheet(ws0)

    ws1 = wb.create_sheet()
    prem_rows = _build_premises_sheet(ws1, specs, fuel_price=fuel_price, fuel_price_meta=fuel_price_meta)

    ws2 = wb.create_sheet()
    _build_steps_sheet(ws2, vehicle, params, prem_rows)

    ws3 = wb.create_sheet()
    _build_comparison_sheet(ws3, result, vehicle, params, prem_rows)

    ws4 = wb.create_sheet()
    _build_sensitivity_sheet(ws4, vehicle, params, prem_rows)

    ws5 = wb.create_sheet()
    _build_scale_sheet(ws5, fleet_size)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
