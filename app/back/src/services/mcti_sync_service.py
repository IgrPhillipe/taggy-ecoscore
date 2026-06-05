"""
Sincroniza fatores de emissão de combustível com a planilha oficial do Programa
Brasileiro GHG Protocol (FGV EAESP).

Nota: O MCTI/SIRENE publica fatores de emissão do SIN (eletricidade), não de
combustíveis líquidos. Os fatores de combustíveis (gasolina, diesel, etanol) vêm
da Ferramenta de Cálculo do Programa GHG Protocol (FGV EAESP).

URL configurável via env var MCTI_EMISSION_FACTORS_URL.
Estrutura esperada: planilha Excel com coluna de nome do combustível e coluna
de fator de emissão CO₂ (kg CO₂/L). O parser busca por palavras-chave nas células
para localizar as linhas de diesel, gasolina e etanol.
"""

import asyncio
import io
import logging
import os
from typing import Any

import httpx
import openpyxl

from src.repositories.technical_specs_repository import TechnicalSpecsRepository

logger = logging.getLogger(__name__)

# URL padrão: Ferramenta de Cálculo do Programa Brasileiro GHG Protocol v2026.0.1 (FGV EAESP).
# Contém fatores de emissão CO₂, CH4 e N2O para gasolina, diesel e etanol.
# Substituível via env var MCTI_EMISSION_FACTORS_URL.
# Página do programa: https://eaesp.fgv.br/centros/centro-estudos-sustentabilidade/projetos/programa-brasileiro-ghg-protocol
_DEFAULT_MCTI_URL = (
    "https://eaesp.fgv.br/sites/eaesp.fgv.br/files/u1087"
    "/ferramenta_ghg_protocol_v2026.0.1.xlsx"
)

# Mapeamento de palavras-chave para nome interno do combustível.
# Mais específico primeiro: "diesel s10"/"s-10" antes de "diesel" genérico,
# "etanol hidratado" antes de "etanol anidro" (hidratado = combustível puro usado nos veículos).
_FUEL_KEYWORDS: dict[str, str] = {
    "diesel s10": "diesel_s10",
    "diesel s-10": "diesel_s10",
    "óleo diesel (puro)": "diesel_s10",
    "oleo diesel (puro)": "diesel_s10",
    "diesel": "diesel_s10",
    "gasolina automotiva (pura)": "gasolina_c",
    "gasolina automotiva": "gasolina_c",
    "gasolina": "gasolina_c",
    "etanol hidratado": "etanol",
    "etanol": "etanol",
    "álcool": "etanol",
    "alcool": "etanol",
}

# Substrings que desqualificam um match de diesel (evita S500, S1800, biodiesel, blend, etc.)
_DIESEL_DISQUALIFIERS = ("s500", "s-500", "s1800", "s-1800", "biodiesel", "b100", "comercial", "blend", "average")

# Substrings que desqualificam matches de etanol (evita "biometanol", "etanol anidro" antes de hidratado, etc.)
_ETANOL_DISQUALIFIERS = ("bio", "anidro")

# Substrings que desqualificam matches de gasolina
_GASOLINA_DISQUALIFIERS = ("aviacao", "aviação", "comercial")

# Sheets prioritárias para busca — contêm fatores de emissão por litro limpos.
# O parser tentará estas abas primeiro antes de varrer todas.
_PRIORITY_SHEETS = (
    "Combustão estacionária",
    "Combustao estacionaria",
    "Fatores de Emissão",
    "Fatores de Emissao",
    "Combustão móvel",
    "Combustao movel",
)

# Range plausível para fatores de emissão em kg CO₂/L.
# Valores reais: etanol ~0.44, gasolina ~2.21, diesel ~2.51.
# Fora deste range indica erro de unidade ou célula errada.
_VALID_FACTOR_MIN = 0.05
_VALID_FACTOR_MAX = 5.0


def _get_mcti_url() -> str:
    return os.environ.get("MCTI_EMISSION_FACTORS_URL", _DEFAULT_MCTI_URL)


def _parse_emission_factors(content: bytes) -> dict[str, float]:
    """
    Parseia o Excel e retorna {combustível: fator_kg_co2_por_litro}.

    Estratégia:
    1. Tenta sheets prioritárias (ex: "Combustão estacionária", "Fatores de Emissão")
       antes de varrer todas — evita falsos positivos em sheets de listas.
    2. Para cada linha, busca keyword na célula e extrai o primeiro número
       no range válido (0.05–5.0) nas células APÓS a keyword.
    3. Disqualifiers por combustível evitam matches errados (blend, anidro, etc.).
    """
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    found: dict[str, float] = {}

    # Ordenar sheets: prioritárias primeiro, depois o restante
    sheet_names = wb.sheetnames
    priority = [n for n in sheet_names if n in _PRIORITY_SHEETS]
    rest = [n for n in sheet_names if n not in _PRIORITY_SHEETS]
    ordered = priority + rest

    for sheet_name in ordered:
        ws = wb[sheet_name]
        for row in ws.iter_rows(values_only=True):
            cells = list(row)
            for i, cell in enumerate(cells):
                if cell is None:
                    continue
                cell_text = str(cell).lower().strip()
                for keyword, fuel_key in _FUEL_KEYWORDS.items():
                    if keyword not in cell_text or fuel_key in found:
                        continue
                    if fuel_key == "diesel_s10" and any(d in cell_text for d in _DIESEL_DISQUALIFIERS):
                        continue
                    if fuel_key == "etanol" and any(d in cell_text for d in _ETANOL_DISQUALIFIERS):
                        continue
                    if fuel_key == "gasolina_c" and any(d in cell_text for d in _GASOLINA_DISQUALIFIERS):
                        continue
                    nums_after = [
                        float(c)
                        for c in cells[i + 1 :]
                        if c is not None
                        and isinstance(c, (int, float))
                        and _VALID_FACTOR_MIN < float(c) < _VALID_FACTOR_MAX
                    ]
                    if nums_after:
                        found[fuel_key] = nums_after[0]
                        break
            if len(found) == 3:
                break
        if len(found) == 3:
            break

    wb.close()
    return found


async def _download_excel(url: str) -> bytes:
    if url.startswith("file://"):
        path = url[len("file://"):]
        try:
            with open(path, "rb") as f:
                return f.read()
        except OSError as e:
            raise RuntimeError(f"Não foi possível ler arquivo local MCTI ({path}): {e}") from e

    async with httpx.AsyncClient(timeout=30.0, follow_redirects=True) as client:
        response = await client.get(url)
        response.raise_for_status()
        content_type = response.headers.get("content-type", "")
        if "html" in content_type or response.content[:2] != b"PK":
            raise RuntimeError(
                f"URL MCTI retornou conteúdo inválido (esperado .xlsx, got '{content_type}'). "
                "O arquivo pode ter mudado de URL — configure MCTI_EMISSION_FACTORS_URL."
            )
        return response.content


async def sync_emission_factors_from_mcti(db: Any) -> dict[str, Any]:
    """
    Baixa a planilha do MCTI/SIRENE, parseia os fatores de emissão e
    persiste no technical_specs (id=1).

    Retorna dict com os campos atualizados e a URL de origem.
    """
    url = _get_mcti_url()
    logger.info("Sincronizando fatores MCTI de: %s", url)

    try:
        content = await _download_excel(url)
    except RuntimeError:
        raise
    except httpx.HTTPError as e:
        raise RuntimeError(f"Falha ao baixar planilha MCTI ({url}): {e}") from e

    try:
        factors = await asyncio.get_event_loop().run_in_executor(
            None, _parse_emission_factors, content
        )
    except Exception as e:
        raise ValueError(f"Falha ao parsear planilha MCTI: {e}") from e

    missing = {"diesel_s10", "gasolina_c", "etanol"} - set(factors.keys())
    if missing:
        raise ValueError(
            f"Fatores não encontrados na planilha MCTI: {missing}. "
            "Verifique a estrutura do arquivo ou configure MCTI_EMISSION_FACTORS_URL."
        )

    payload = {
        "emission_factor_diesel_s10": factors["diesel_s10"],
        "emission_factor_gasolina_c": factors["gasolina_c"],
        "emission_factor_etanol": factors["etanol"],
    }

    repo = TechnicalSpecsRepository(db)
    await repo.upsert_by_id(id=1, data=payload)
    await db.commit()

    logger.info("Fatores MCTI atualizados: %s", payload)
    return {"updated": payload, "source_url": url}
